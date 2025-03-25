import csv
import time
import os
from openpyxl import load_workbook

# this code is used to format and standardise most excel sheets since the naming convention/format of header and isbn13 is not standardised

def normalize_header(header_name):
    return header_name.replace("-", "").replace("_", "").replace(" ", "").lower()

def format_isbn(isbn):
    # setting the format for all isbn to be xxx-xxxxxxxxxx
    isbn = ''.join(c for c in isbn if c.isdigit())
    return isbn[:3] + "-" + isbn[3:] 

def read_csv(inputname):
    with open(inputname, mode="r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file)
        header = next(reader)
        rows = list(reader)
    return header, rows

def read_xlsx(inputname):
    workbook = load_workbook(filename=inputname)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
    return header, rows

def write_csv(outputname, header, rows):
    with open(outputname, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(header)
        writer.writerows(rows)

def isbnformat(inputname, outputname):
    start_time = time.perf_counter()

    file_extension = os.path.splitext(inputname)[1].lower()
    if file_extension == ".csv":
        header, rows = read_csv(inputname)
    elif file_extension == ".xlsx":
        header, rows = read_xlsx(inputname)
    else:
        raise ValueError("Unsupported file format. Only .csv and .xlsx are supported.")

    normalized_header = [normalize_header(col) for col in header]
    
    # identify ISBN-13 column and rename it to 'isbn13'
    isbn_index = next((i for i, col in enumerate(normalized_header) if "isbn" in col and "13" in col), -1)
    if isbn_index == -1:
        raise ValueError("No ISBN-13 column found in the file.")
    
    header[isbn_index] = "isbn13"

    for row in rows:
        if len(row) > isbn_index and row[isbn_index]:
            row[isbn_index] = format_isbn(str(row[isbn_index]))

    # for ease of use, output is currently always a csv
    write_csv(outputname, header, rows)

    end_time = time.perf_counter()
    elapsed_time = end_time - start_time

    print(f"Format script executed in {elapsed_time:.2f} seconds")
    print(f"Formatted data saved to {outputname}")
